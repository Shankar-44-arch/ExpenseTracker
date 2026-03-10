from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from .models import Transaction, Category, BudgetAlert
from django.db.models import Sum, Q
from django.utils import timezone
from collections import defaultdict
from datetime import timedelta
import csv
import io
import json
import base64


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password.')

    return render(request, 'loginmain.html')


def signup_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')

        if not username or not email or not password1:
            messages.error(request, 'All fields are required.')
        elif password1 != password2:
            messages.error(request, 'Passwords do not match.')
        elif len(password1) < 8:
            messages.error(request, 'Password must be at least 8 characters.')
        elif User.objects.filter(username=username).exists():
            messages.error(request, 'Username already taken.')
        elif User.objects.filter(email=email).exists():
            messages.error(request, 'Email already registered.')
        else:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password1,
            )
            # Create default categories for the new user
            ensure_default_categories(user)

            login(request, user)
            messages.success(request, 'Account created successfully!')
            return redirect('dashboard')

    return render(request, 'loginmain.html')


def logout_view(request):
    logout(request)
    return redirect('login')


DEFAULT_CATEGORIES = [
    ('🍔', 'Food', 'Groceries, dining out, and snacks'),
    ('🚕', 'Travel', 'Taxi fares, fuel, and public transport'),
    ('💡', 'Electricity Bill', 'Electricity and power bills'),
    ('🏋️', 'Gym', 'Gym memberships and fitness expenses'),
    ('📚', 'Education', 'Tuition, books, and courses'),
    ('🛒', 'Shopping', 'Clothes, accessories, and online shopping'),
    ('🏠', 'Housing', 'Rent, mortgage, and home maintenance'),
    ('🎬', 'Entertainment', 'Movies, streaming, and hobbies'),
    ('🏥', 'Health', 'Medical checkups, medicines, and insurance'),
    ('📱', 'Subscriptions', 'Phone, internet, and app subscriptions'),
]


def ensure_default_categories(user):
    """Create default categories for a user if they have none."""
    if not Category.objects.filter(user=user).exists():
        Category.objects.bulk_create([
            Category(user=user, icon=icon, name=name, description=desc)
            for icon, name, desc in DEFAULT_CATEGORIES
        ])


@login_required
def dashboard_view(request):
    user = request.user
    ensure_default_categories(user)
    today = timezone.now().date()
    transactions = Transaction.objects.filter(user=user)

    total_income = transactions.filter(transaction_type='income').aggregate(s=Sum('amount'))['s'] or 0
    total_expense = transactions.filter(transaction_type='expense').aggregate(s=Sum('amount'))['s'] or 0
    savings = total_income - total_expense

    # Recent transactions (last 5)
    recent = transactions[:5]

    # Monthly trend — last 6 months expenses
    month_labels = []
    month_expense_data = []
    month_income_data = []
    for i in range(5, -1, -1):
        d = today.replace(day=1) - timedelta(days=i * 28)
        month_labels.append(d.strftime('%b'))
        m_txns = transactions.filter(date__year=d.year, date__month=d.month)
        month_expense_data.append(float(m_txns.filter(transaction_type='expense').aggregate(s=Sum('amount'))['s'] or 0))
        month_income_data.append(float(m_txns.filter(transaction_type='income').aggregate(s=Sum('amount'))['s'] or 0))

    # Expense by category
    cat_totals = (
        transactions.filter(transaction_type='expense')
        .values('category')
        .annotate(total=Sum('amount'))
        .order_by('-total')
    )
    cat_labels = [c['category'] for c in cat_totals]
    cat_data = [float(c['total']) for c in cat_totals]

    # Budget alerts — check current month spending vs limits
    now = timezone.now().date()
    alerts = BudgetAlert.objects.filter(user=user)
    budget_warnings = []
    for alert in alerts:
        month_spent = (
            transactions.filter(
                transaction_type='expense',
                category=alert.category,
                date__year=now.year,
                date__month=now.month,
            ).aggregate(s=Sum('amount'))['s'] or 0
        )
        if month_spent >= alert.limit_amount:
            budget_warnings.append({
                'category': alert.category,
                'limit': float(alert.limit_amount),
                'spent': float(month_spent),
                'over': float(month_spent - alert.limit_amount),
            })

    context = {
        'total_income': total_income,
        'total_expense': total_expense,
        'savings': savings,
        'recent': recent,
        'budget_warnings': budget_warnings,
        'month_labels': json.dumps(month_labels),
        'month_expense_data': json.dumps(month_expense_data),
        'month_income_data': json.dumps(month_income_data),
        'cat_labels': json.dumps(cat_labels),
        'cat_data': json.dumps(cat_data),
    }
    return render(request, 'dash.html', context)


@login_required
def transaction_view(request):
    user = request.user
    transactions = Transaction.objects.filter(user=user)
    categories = Category.objects.filter(user=user)

    # Filters
    search = request.GET.get('search', '').strip()
    cat_filter = request.GET.get('category', '')
    time_filter = request.GET.get('time', '')

    if search:
        transactions = transactions.filter(
            Q(description__icontains=search) | Q(category__icontains=search)
        )
    if cat_filter:
        transactions = transactions.filter(category=cat_filter)
    if time_filter == 'today':
        transactions = transactions.filter(date=timezone.now().date())
    elif time_filter == 'month':
        now = timezone.now().date()
        transactions = transactions.filter(date__year=now.year, date__month=now.month)

    context = {
        'transactions': transactions,
        'categories': categories,
        'search': search,
        'cat_filter': cat_filter,
        'time_filter': time_filter,
    }
    return render(request, 'transaction.html', context)


@login_required
def category_view(request):
    categories = Category.objects.filter(user=request.user)
    return render(request, 'cat.html', {'categories': categories})

@login_required
def report_view(request):
    user = request.user
    today = timezone.now().date()
    transactions = Transaction.objects.filter(user=user)

    # Monthly trend — last 6 months
    month_labels = []
    month_expense_data = []
    month_income_data = []
    for i in range(5, -1, -1):
        d = today.replace(day=1) - timedelta(days=i * 28)
        month_labels.append(d.strftime('%b'))
        m_txns = transactions.filter(date__year=d.year, date__month=d.month)
        month_expense_data.append(float(m_txns.filter(transaction_type='expense').aggregate(s=Sum('amount'))['s'] or 0))
        month_income_data.append(float(m_txns.filter(transaction_type='income').aggregate(s=Sum('amount'))['s'] or 0))

    # Expense by category
    cat_totals = (
        transactions.filter(transaction_type='expense')
        .values('category')
        .annotate(total=Sum('amount'))
        .order_by('-total')
    )
    cat_labels = [c['category'] for c in cat_totals]
    cat_data = [float(c['total']) for c in cat_totals]

    context = {
        'month_labels': json.dumps(month_labels),
        'month_expense_data': json.dumps(month_expense_data),
        'month_income_data': json.dumps(month_income_data),
        'cat_labels': json.dumps(cat_labels),
        'cat_data': json.dumps(cat_data),
    }
    return render(request, 'report.html', context)

@login_required
def calculator_view(request):
    return render(request, 'cal.html')

@login_required
def setting_view(request):
    return render(request, 'settings.html')


# ================= TRANSACTION ACTIONS =================

@login_required
@require_POST
def add_transaction(request):
    user = request.user
    description = request.POST.get('description', '').strip()
    category = request.POST.get('category', '').strip()
    amount = request.POST.get('amount', '')
    date = request.POST.get('date', '')
    txn_type = request.POST.get('transaction_type', 'expense')

    if not description or not amount or not date:
        messages.error(request, 'All fields are required.')
        return redirect('transaction')

    if txn_type == 'income':
        category = 'Income'

    Transaction.objects.create(
        user=user,
        description=description,
        category=category,
        amount=amount,
        transaction_type=txn_type,
        date=date,
    )
    messages.success(request, 'Transaction added.')
    return redirect('transaction')


@login_required
@require_POST
def edit_transaction(request, pk):
    txn = get_object_or_404(Transaction, pk=pk, user=request.user)
    txn.description = request.POST.get('description', txn.description).strip()
    txn.category = request.POST.get('category', txn.category).strip()
    txn.amount = request.POST.get('amount', txn.amount)
    txn.date = request.POST.get('date', txn.date)
    txn.transaction_type = request.POST.get('transaction_type', txn.transaction_type)
    if txn.transaction_type == 'income':
        txn.category = 'Income'
    txn.save()
    messages.success(request, 'Transaction updated.')
    return redirect('transaction')


@login_required
@require_POST
def delete_transaction(request, pk):
    txn = get_object_or_404(Transaction, pk=pk, user=request.user)
    txn.delete()
    messages.success(request, 'Transaction deleted.')
    return redirect('transaction')


# ================= CATEGORY ACTIONS =================

@login_required
@require_POST
def add_category(request):
    name = request.POST.get('name', '').strip()
    icon = request.POST.get('icon', '💰').strip()
    description = request.POST.get('description', '').strip()

    if not name:
        messages.error(request, 'Category name is required.')
        return redirect('category')

    if Category.objects.filter(user=request.user, name=name).exists():
        messages.error(request, 'Category already exists.')
        return redirect('category')

    Category.objects.create(
        user=request.user, name=name, icon=icon, description=description
    )
    messages.success(request, f'Category "{name}" created.')
    return redirect('category')


@login_required
@require_POST
def edit_category(request, pk):
    cat = get_object_or_404(Category, pk=pk, user=request.user)
    name = request.POST.get('name', '').strip()
    icon = request.POST.get('icon', cat.icon).strip()
    description = request.POST.get('description', '').strip()

    if not name:
        messages.error(request, 'Category name is required.')
        return redirect('category')

    if Category.objects.filter(user=request.user, name=name).exclude(pk=pk).exists():
        messages.error(request, 'A category with that name already exists.')
        return redirect('category')

    old_name = cat.name
    cat.name = name
    cat.icon = icon
    cat.description = description
    cat.save()

    # Update category name in existing transactions
    if old_name != name:
        Transaction.objects.filter(user=request.user, category=old_name).update(category=name)

    messages.success(request, f'Category "{name}" updated.')
    return redirect('category')


@login_required
@require_POST
def delete_category(request, pk):
    cat = get_object_or_404(Category, pk=pk, user=request.user)
    cat.delete()
    messages.success(request, 'Category deleted.')
    return redirect('category')


# ================= SETTINGS ACTIONS =================

@login_required
@require_POST
def update_profile(request):
    user = request.user
    username = request.POST.get('username', '').strip()
    email = request.POST.get('email', '').strip()

    if not username or not email:
        messages.error(request, 'Username and email are required.')
        return redirect('setting')

    if User.objects.filter(username=username).exclude(pk=user.pk).exists():
        messages.error(request, 'That username is already taken.')
        return redirect('setting')

    if User.objects.filter(email=email).exclude(pk=user.pk).exists():
        messages.error(request, 'That email is already in use.')
        return redirect('setting')

    user.username = username
    user.email = email

    first_name = request.POST.get('first_name', '').strip()
    last_name = request.POST.get('last_name', '').strip()
    if first_name:
        user.first_name = first_name
    if last_name:
        user.last_name = last_name

    user.save()
    messages.success(request, 'Profile updated successfully.')
    return redirect('setting')


@login_required
@require_POST
def change_password(request):
    user = request.user
    current = request.POST.get('current_password', '')
    new1 = request.POST.get('new_password', '')
    new2 = request.POST.get('confirm_password', '')

    if not user.check_password(current):
        messages.error(request, 'Current password is incorrect.')
        return redirect('setting')

    if new1 != new2:
        messages.error(request, 'New passwords do not match.')
        return redirect('setting')

    if len(new1) < 8:
        messages.error(request, 'New password must be at least 8 characters.')
        return redirect('setting')

    user.set_password(new1)
    user.save()
    update_session_auth_hash(request, user)
    messages.success(request, 'Password changed successfully.')
    return redirect('setting')


@login_required
@require_POST
def delete_account(request):
    user = request.user
    password = request.POST.get('password', '')

    if not user.check_password(password):
        messages.error(request, 'Incorrect password. Account not deleted.')
        return redirect('setting')

    user.delete()
    messages.success(request, 'Account deleted.')
    return redirect('login')


# ================= REPORT EXPORTS =================

# ================= ALERT ACTIONS =================

@login_required
def alert_view(request):
    user = request.user
    categories = Category.objects.filter(user=user)
    alerts = BudgetAlert.objects.filter(user=user)
    now = timezone.now().date()

    alert_list = []
    for alert in alerts:
        spent = (
            Transaction.objects.filter(
                user=user,
                transaction_type='expense',
                category=alert.category,
                date__year=now.year,
                date__month=now.month,
            ).aggregate(s=Sum('amount'))['s'] or 0
        )
        percent = min(int(float(spent) / float(alert.limit_amount) * 100), 100) if alert.limit_amount else 0
        alert_list.append({
            'obj': alert,
            'spent': float(spent),
            'percent': percent,
            'exceeded': float(spent) >= float(alert.limit_amount),
        })

    return render(request, 'alert.html', {'alert_list': alert_list, 'categories': categories})


@login_required
@require_POST
def add_alert(request):
    category = request.POST.get('category', '').strip()
    limit_amount = request.POST.get('limit_amount', '')

    if not category or not limit_amount:
        messages.error(request, 'Category and limit are required.')
        return redirect('alert')

    if BudgetAlert.objects.filter(user=request.user, category=category).exists():
        messages.error(request, f'Alert for "{category}" already exists.')
        return redirect('alert')

    BudgetAlert.objects.create(user=request.user, category=category, limit_amount=limit_amount)
    messages.success(request, f'Budget alert for "{category}" set.')
    return redirect('alert')


@login_required
@require_POST
def edit_alert(request, pk):
    alert = get_object_or_404(BudgetAlert, pk=pk, user=request.user)
    limit_amount = request.POST.get('limit_amount', '')

    if not limit_amount:
        messages.error(request, 'Limit amount is required.')
        return redirect('alert')

    alert.limit_amount = limit_amount
    alert.save()
    messages.success(request, f'Alert for "{alert.category}" updated.')
    return redirect('alert')


@login_required
@require_POST
def delete_alert(request, pk):
    alert = get_object_or_404(BudgetAlert, pk=pk, user=request.user)
    alert.delete()
    messages.success(request, 'Budget alert removed.')
    return redirect('alert')


@login_required
def export_csv(request):
    transactions = Transaction.objects.filter(user=request.user)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="transactions_report.csv"'

    writer = csv.writer(response)
    writer.writerow(['Description', 'Category', 'Type', 'Amount', 'Date'])
    for t in transactions:
        writer.writerow([t.description, t.category, t.transaction_type, t.amount, t.date.strftime('%d/%m/%Y')])

    return response


@login_required
def export_excel(request):
    try:
        import openpyxl
    except ImportError:
        return HttpResponse('openpyxl is not installed. Run: pip install openpyxl', status=500)

    transactions = Transaction.objects.filter(user=request.user)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'

    headers = ['Description', 'Category', 'Type', 'Amount', 'Date']
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')

    for t in transactions:
        ws.append([t.description, t.category, t.transaction_type, float(t.amount), t.date.strftime('%d/%m/%Y')])

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="transactions_report.xlsx"'
    return response


@login_required
def export_pdf(request):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        return HttpResponse('reportlab is not installed. Run: pip install reportlab', status=500)

    transactions = Transaction.objects.filter(user=request.user)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()
    elements.append(Paragraph('Transaction Report', styles['Title']))
    elements.append(Spacer(1, 20))

    data = [['Description', 'Category', 'Type', 'Amount', 'Date']]
    for t in transactions:
        data.append([t.description, t.category, t.transaction_type, str(t.amount), t.date.strftime('%d/%m/%Y')])

    if len(data) == 1:
        data.append(['No transactions found', '', '', '', ''])

    table = Table(data, colWidths=[140, 90, 70, 80, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9f9f9')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f2f2')]),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="transactions_report.pdf"'
    return response


@login_required
def export_image(request):
    """Return chart data as JSON for client-side canvas-to-image download."""
    transactions = Transaction.objects.filter(user=request.user)

    category_totals = {}
    for t in transactions:
        category_totals[t.category] = category_totals.get(t.category, 0) + float(t.amount)

    return JsonResponse({
        'labels': list(category_totals.keys()) or ['No Data'],
        'data': list(category_totals.values()) or [0],
    })




